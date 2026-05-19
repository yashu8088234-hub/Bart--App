import streamlit as st
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
from st_aggrid import AgGrid, GridOptionsBuilder
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import time

# ========================================================
# PAGE CONFIG
# =========================================================

st.set_page_config(layout="wide", page_title="Stock Overview")
st.title("📦 BART - Stock Management (All Branches)")

# =========================================================
# ERROR
# =========================================================

def show_api_error():
    st.error("⚠️ API Error")
    st.stop()

# =========================================================
# GOOGLE AUTH
# =========================================================

creds_dict = st.secrets["GOOGLE_CREDS_JSON"]

scope = [
    "https://spreadsheets.google.com/feeds",
    "https://www.googleapis.com/auth/drive"
]

@st.cache_resource
def get_client():
    creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
    return gspread.authorize(creds)

client = get_client()

# =========================================================
# BRANCHES (FORCE ALL)
# =========================================================

@st.cache_data(ttl=600)
def load_branches():
    sheet = client.open("MASTERBRANCHSHEET").sheet1
    data = sheet.get_all_records()

    branches = []

    for b in data:

        if b.get("SheetID") and b.get("BranchName"):

            branches.append({
                "BranchName": str(b["BranchName"]).strip(),
                "SheetID": str(b["SheetID"]).strip()
            })

    return branches

branches = load_branches()

branch_names = [b["BranchName"] for b in branches]

# =========================================================
# FETCH SHEETS (ROBUST RETRY)
# =========================================================

MAX_RETRIES = 10
RETRY_DELAY = 60

branch_cache = {}

def fetch_branch(branch):

    branch_name = branch["BranchName"]

    try:

        ws = client.open_by_key(branch["SheetID"]).worksheet("Stocks")

        # OPTIONAL OPTIMIZATION
        # data = ws.get("A1:AZ500")

        data = ws.get_all_values()

        branch_cache[branch_name] = data

        return {
            "branch": branch_name,
            "success": True,
            "data": data,
            "error": None
        }

    except Exception as e:

        if branch_name in branch_cache:

            return {
                "branch": branch_name,
                "success": False,
                "data": branch_cache[branch_name],
                "error": str(e)
            }

        return {
            "branch": branch_name,
            "success": False,
            "data": [],
            "error": str(e)
        }

@st.cache_data(ttl=600)
def load_all_data(branches):

    completed = {}
    failed = []

    progress_bar = st.progress(0)
    status_text = st.empty()

    # =====================================================
    # INITIAL LOAD
    # =====================================================

    with ThreadPoolExecutor(max_workers=3) as ex:

        futures = {
            ex.submit(fetch_branch, b): b
            for b in branches
        }

        completed_count = 0

        for future in as_completed(futures):

            result = future.result()

            branch_name = result["branch"]

            if result["success"] or result["data"]:

                completed[branch_name] = result["data"]

            else:

                failed.append(futures[future])

            completed_count += 1

            progress_bar.progress(completed_count / len(branches))

            status_text.info(
                f"Loaded {completed_count}/{len(branches)} branches"
            )

    # =====================================================
    # RETRY FAILED BRANCHES
    # =====================================================

    retry_round = 1

    while failed and retry_round <= MAX_RETRIES:

        failed_names = [b["BranchName"] for b in failed]

        st.warning(
            f"""
Retry Round {retry_round}/{MAX_RETRIES}

Failed Branches:
{', '.join(failed_names)}

Retrying in {RETRY_DELAY} seconds...
"""
        )

        time.sleep(RETRY_DELAY)

        retry_failed = []

        with ThreadPoolExecutor(max_workers=3) as ex:

            futures = {
                ex.submit(fetch_branch, b): b
                for b in failed
            }

            for future in as_completed(futures):

                result = future.result()

                branch_name = result["branch"]

                if result["success"] or result["data"]:

                    completed[branch_name] = result["data"]

                    st.success(f"✅ Recovered: {branch_name}")

                else:

                    retry_failed.append(futures[future])

                    st.error(
                        f"❌ Failed: {branch_name}"
                    )

        failed = retry_failed

        retry_round += 1

    # =====================================================
    # FINAL STATUS
    # =====================================================

    if failed:

        failed_names = [b["BranchName"] for b in failed]

        st.error(
            f"""
Some branches failed after {MAX_RETRIES} retries:

{', '.join(failed_names)}
"""
        )

    else:

        st.success("✅ All branches loaded successfully")

    # maintain original order

    ordered = []

    for b in branches:

        ordered.append(
            (
                b["BranchName"],
                completed.get(b["BranchName"], [])
            )
        )

    return ordered

all_data = load_all_data(branches)

# =========================================================
# REFRESH
# =========================================================

if st.button("🔄 Refresh Data"):

    st.cache_data.clear()

    st.cache_resource.clear()

    branch_cache.clear()

    st.rerun()

# =========================================================
# DATE
# =========================================================

selected_date = st.date_input("📅 Select Date")

selected_date_str = selected_date.strftime("%Y-%m-%d")

# =========================================================
# PROCESS
# =========================================================

@st.cache_data(ttl=600)
def process_stock(all_data, selected_date_str, branch_names):

    daily = {}

    weekly = {}

    for branch_name, raw in all_data:

        if not raw or len(raw) < 2:
            continue

        headers = [str(x).strip() for x in raw[0]]

        date_index = None

        for i, h in enumerate(headers):

            if h == selected_date_str:

                date_index = i

                break

        current_section = None

        for row in raw:

            if not row:
                continue

            text = " ".join([str(x) for x in row]).lower()

            if "daily item" in text:

                current_section = "daily"

                continue

            if "weekly item" in text:

                current_section = "weekly"

                continue

            if current_section is None:
                continue

            item = str(row[0]).strip() if len(row) > 0 else ""

            sku = str(row[1]).strip() if len(row) > 1 else ""

            uom = str(row[2]).strip() if len(row) > 2 else ""

            if not item:
                continue

            key = f"{item}_{sku}_{uom}"

            target = daily if current_section == "daily" else weekly

            if key not in target:

                target[key] = {
                    "Item Name": item,
                    "SKU": sku,
                    "UOM": uom
                }

                for bn in branch_names:

                    target[key][bn] = 0

            qty = 0

            try:

                if date_index is not None and len(row) > date_index:

                    val = row[date_index]

                    qty = 0 if val in ["", None] else float(val)

            except:

                qty = 0

            target[key][branch_name] = qty

    return daily, weekly

daily_items, weekly_items = process_stock(
    all_data,
    selected_date_str,
    branch_names
)

# =========================================================
# DF
# =========================================================

def build_df(data_dict):

    rows = []

    for _, v in data_dict.items():

        row = {
            "Item Name": v["Item Name"],
            "SKU": v["SKU"],
            "UOM": v["UOM"]
        }

        for b in branch_names:

            row[b] = v.get(b, 0)

        rows.append(row)

    return pd.DataFrame(rows)

daily_df = build_df(daily_items)

weekly_df = build_df(weekly_items)

# =========================================================
# GRID
# =========================================================

def get_width(series, min_width):

    try:

        series = series.fillna("").astype(str)

        return max(
            min_width,
            int(series.map(len).max() * 6)
        )

    except:

        return min_width

def render_grid(df, title):

    st.subheader(title)

    if df is None or df.empty:

        st.warning("No Data")

        return

    gb = GridOptionsBuilder.from_dataframe(df)

    gb.configure_column(
        "Item Name",
        pinned="left",
        minWidth=get_width(df["Item Name"], 120)
    )

    gb.configure_column(
        "SKU",
        pinned="left",
        minWidth=get_width(df["SKU"], 80)
    )

    gb.configure_column(
        "UOM",
        pinned="left",
        minWidth=get_width(df["UOM"], 80)
    )

    for col in branch_names:

        if col in df.columns:

            gb.configure_column(col, minWidth=120)

    gb.configure_default_column(
        resizable=True,
        sortable=True,
        filter=True
    )

    AgGrid(
        df,
        gridOptions=gb.build(),
        theme="streamlit",
        fit_columns_on_grid_load=True,
        key=title
    )

render_grid(daily_df, "📦 Daily Items Stock")

render_grid(weekly_df, "📦 Weekly Items Stock")

# =========================================================
# EXCEL
# =========================================================

def create_excel(daily_df, weekly_df):

    output = BytesIO()

    wb = Workbook()

    ws = wb.active

    ws.title = "Stock"

    header_font = Font(bold=True)

    align = Alignment(
        horizontal="center",
        vertical="center"
    )

    zebra = PatternFill(
        "solid",
        fgColor="F2F2F2"
    )

    def write(title, df, start):

        rows = list(
            dataframe_to_rows(
                df,
                index=False,
                header=True
            )
        )

        if not rows:
            return start + 2

        cols = len(rows[0])

        ws.merge_cells(
            start_row=start,
            start_column=1,
            end_row=start,
            end_column=cols
        )

        ws.cell(
            row=start,
            column=1,
            value=title
        ).font = Font(
            bold=True,
            size=14
        )

        r0 = start + 2

        for r_i, row in enumerate(rows):

            for c_i, val in enumerate(row, 1):

                cell = ws.cell(
                    row=r0 + r_i,
                    column=c_i,
                    value=val
                )

                cell.alignment = align

                if r_i == 0:

                    cell.font = header_font

                elif r_i % 2 == 0:

                    cell.fill = zebra

        return r0 + len(rows) + 3

    n = write("DAILY", daily_df, 1)

    write("WEEKLY", weekly_df, n)

    for col in ws.columns:

        try:

            letter = get_column_letter(col[0].column)

            ws.column_dimensions[letter].width = max(
                len(str(c.value or ""))
                for c in col
            ) + 3

        except:
            pass

    wb.save(output)

    output.seek(0)

    return output

excel_file = create_excel(
    daily_df,
    weekly_df
)

st.download_button(
    "📥 Download Excel",
    excel_file,
    file_name="stock_report.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
